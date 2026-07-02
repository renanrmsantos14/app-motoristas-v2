import csv
import datetime as dt
import json
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path.home() / "Downloads"
TRACKER_FILES = [
    DOWNLOADS / "b0edc78d-264b-4f3a-9782-829abfc80e3f.xlsx",
    DOWNLOADS / "c03e3c8e-2325-40ab-9ccd-8b68334afbeb.xlsx",
]
AGENDA_FILE = next(DOWNLOADS.glob("*Agenda*01-07-2026 17-18-47.xlsx"))
OUTPUT_JSON = ROOT / "outputs" / "service-vehicle-tracker-corrections-2026-07-01.json"
OUTPUT_CSV = ROOT / "outputs" / "service-vehicle-tracker-service-scope-2026-07-01.csv"
OUTPUT_SCRIPT = ROOT / "scripts" / "update-service-vehicles-from-tracker.console.js"


def normalize_text(value):
    text = str(value or "").strip().lower()
    text = "".join(
        char for char in unicodedata.normalize("NFD", text)
        if unicodedata.category(char) != "Mn"
    )
    return " ".join(text.split())


def parse_service_datetime(value):
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time())
    if value is None:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            return dt.datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def parse_tracker_datetime(day_value, time_value):
    if isinstance(day_value, dt.datetime):
        day = day_value.date()
    elif isinstance(day_value, dt.date):
        day = day_value
    else:
        try:
            day = dt.datetime.strptime(str(day_value).strip(), "%d/%m/%Y").date()
        except ValueError:
            return None

    if isinstance(time_value, dt.datetime):
        time = time_value.time()
    elif isinstance(time_value, dt.time):
        time = time_value
    else:
        time = None
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                time = dt.datetime.strptime(str(time_value).strip(), fmt).time()
                break
            except ValueError:
                pass
        if time is None:
            return None
    return dt.datetime.combine(day, time)


def format_datetime(value):
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else ""


def load_tracker_windows():
    events_by_plate = defaultdict(list)
    all_timestamps = []
    event_count = 0

    for path in TRACKER_FILES:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        headers = [normalize_text(value) for value in next(sheet.iter_rows(values_only=True))]
        index = {name: idx for idx, name in enumerate(headers)}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            timestamp = parse_tracker_datetime(row[index["dia"]], row[index["hora"]])
            plate = str(row[index["veiculo"]] or "").strip().upper()
            driver = str(row[index["motorista"]] or "").strip()
            if not timestamp or not plate or not driver:
                continue

            event = {
                "timestamp": timestamp,
                "driver": driver,
                "driverKey": normalize_text(driver),
                "event": str(row[index["evento"]] or "").strip(),
            }
            events_by_plate[plate].append(event)
            all_timestamps.append(timestamp)
            event_count += 1

    if not all_timestamps:
        raise RuntimeError("Nenhum evento valido encontrado no rastreador.")

    coverage_start = min(all_timestamps)
    coverage_end = max(all_timestamps)
    windows_by_driver = defaultdict(list)
    raw_window_count = 0

    for plate, events in events_by_plate.items():
        events.sort(key=lambda item: item["timestamp"])
        merged_for_plate = []

        for idx, event in enumerate(events):
            end = events[idx + 1]["timestamp"] if idx + 1 < len(events) else coverage_end
            if end <= event["timestamp"]:
                continue
            raw_window_count += 1

            current = {
                "plate": plate,
                "start": event["timestamp"],
                "end": end,
                "driverNameFromTracker": event["driver"],
                "driverKey": event["driverKey"],
                "firstEvent": event["event"],
                "lastEvent": event["event"],
            }

            previous = merged_for_plate[-1] if merged_for_plate else None
            if (
                previous
                and previous["driverKey"] == current["driverKey"]
                and previous["end"] == current["start"]
            ):
                previous["end"] = current["end"]
                previous["lastEvent"] = current["lastEvent"]
            else:
                merged_for_plate.append(current)

        for window in merged_for_plate:
            windows_by_driver[window["driverKey"]].append({
                "plate": window["plate"],
                "start": format_datetime(window["start"]),
                "end": format_datetime(window["end"]),
                "driverNameFromTracker": window["driverNameFromTracker"],
                "firstEvent": window["firstEvent"],
                "lastEvent": window["lastEvent"],
            })

    for windows in windows_by_driver.values():
        windows.sort(key=lambda item: item["start"])

    merged_window_count = sum(len(windows) for windows in windows_by_driver.values())
    summary = {
        "trackerEvents": event_count,
        "trackerRawWindows": raw_window_count,
        "trackerMergedWindows": merged_window_count,
        "trackerPlates": len(events_by_plate),
        "trackerDrivers": len(windows_by_driver),
        "trackerCoverageStartLocal": format_datetime(coverage_start),
        "trackerCoverageEndLocal": format_datetime(coverage_end),
    }
    return dict(windows_by_driver), summary


def load_service_scope():
    workbook = openpyxl.load_workbook(AGENDA_FILE, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    headers = [normalize_text(value) for value in next(sheet.iter_rows(values_only=True))]
    index = {name: idx for idx, name in enumerate(headers)}

    required = {
        "id": "id",
        "guid": "(nao modificar) guid reserva de veiculos",
        "status": "status de operacao",
        "start": "data e horario de saida",
        "driver": "motorista",
        "vehicle": "veiculo",
    }
    missing = [label for label in required.values() if label not in index]
    if missing:
        raise RuntimeError(f"Colunas ausentes na agenda: {', '.join(missing)}")

    services = []
    stats = Counter()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        service_id = str(row[index[required["guid"]]] or "").strip().lower()
        service_number = str(row[index[required["id"]]] or "").strip()
        if not service_id or not service_number:
            continue

        start = parse_service_datetime(row[index[required["start"]]])
        services.append({
            "serviceId": service_id,
            "serviceNumber": service_number,
            "statusFromAgenda": str(row[index[required["status"]]] or "").strip(),
            "startLocalFromAgenda": format_datetime(start),
            "driverNameFromAgenda": str(row[index[required["driver"]]] or "").strip(),
            "vehiclePlateFromAgenda": str(row[index[required["vehicle"]]] or "").strip().upper(),
        })
        stats["servicesInAgendaExport"] += 1

    services.sort(key=lambda item: (item["startLocalFromAgenda"], item["serviceNumber"]))
    return services, dict(stats)


def write_json_and_csv(summary, services, tracker_windows_by_driver):
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(
        json.dumps(
            {
                "summary": summary,
                "services": services,
                "trackerWindowsByDriver": tracker_windows_by_driver,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    with OUTPUT_CSV.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=list(services[0].keys()) if services else [])
        writer.writeheader()
        writer.writerows(services)


def write_console_script(services, tracker_windows_by_driver, tracker_summary):
    services_json = json.dumps(services, ensure_ascii=True, separators=(",", ":"))
    windows_json = json.dumps(tracker_windows_by_driver, ensure_ascii=True, separators=(",", ":"))
    coverage_json = json.dumps(
        {
            "start": tracker_summary["trackerCoverageStartLocal"],
            "end": tracker_summary["trackerCoverageEndLocal"],
        },
        ensure_ascii=True,
        separators=(",", ":"),
    )

    script = f'''/**
 * Console script: atualiza veiculo dos servicos pelo rastreador.
 *
 * Correcao importante:
 * - nao confia no motorista do Excel para decidir atualizacao;
 * - consulta todos os servicos pelo GUID via Web API, sem abrir registro um por um;
 * - usa o motorista atual do lookup cr40f_motorista;
 * - calcula a placa pelo rastreador no horario atual do servico;
 * - se o Dataverse esta sem veiculo, preenche com a placa do rastreador;
 * - se o Dataverse ja tem veiculo diferente do rastreador, alerta sem sobrescrever;
 * - se motorista nao existe no rastreador, ou se ha ambiguidade, nao mexe.
 *
 * Uso:
 * 1. Cole no console do Model-driven App no ambiente alvo.
 * 2. Rode uma vez e aguarde terminar. Resultado fica em window.ServiceVehicleCorrectionLastResult.
 * 3. Se quiser simular sem aplicar, mude DRY_RUN para true antes de rodar.
 */
(async () => {{
  const DRY_RUN = false;
  const STRICT_SERVICE_NUMBER = true;
  const PAGE_SIZE = 5000;
  const ACTION_LOG_LIMIT = 1000;
  const PROGRESS_EVERY = 50;

  const SERVICE_ENTITY_SET = "cr40f_reservadeveculoses";
  const VEHICLE_ENTITY_SET = "cr40f_veiculoses";
  const EMPLOYEE_ENTITY_SET = "cr40f_funcionarioses";
  const SERVICE_ID_FIELD = "cr40f_reservadeveculosid";
  const SERVICE_NUMBER_FIELD = "cr40f_id";
  const SERVICE_DATE_FIELD = "cr40f_dataehorriodesada";
  const SERVICE_VEHICLE_LOOKUP = "_cr40f_veiculo_value";
  const SERVICE_DRIVER_LOOKUP = "_cr40f_motorista_value";
  const SERVICE_VEHICLE_NAV = "cr40f_Veiculo";
  const VEHICLE_ID_FIELD = "cr40f_veiculosid";
  const VEHICLE_PLATE_FIELD = "cr40f_placa";
  const EMPLOYEE_ID_FIELD = "cr40f_funcionariosid";
  const EMPLOYEE_NAME_FIELD = "cr40f_nomecompleto";
  const DRIVER_FORMATTED = `${{SERVICE_DRIVER_LOOKUP}}@OData.Community.Display.V1.FormattedValue`;
  const VEHICLE_FORMATTED = `${{SERVICE_VEHICLE_LOOKUP}}@OData.Community.Display.V1.FormattedValue`;
  const DATE_FORMATTED = `${{SERVICE_DATE_FIELD}}@OData.Community.Display.V1.FormattedValue`;

  const SERVICES = {services_json};
  const TRACKER_WINDOWS_BY_DRIVER = {windows_json};
  const TRACKER_COVERAGE = {coverage_json};

  if (!window.Xrm?.Utility?.getGlobalContext) {{
    throw new Error("Xrm.Utility nao encontrado. Abra este script dentro do model-driven app.");
  }}

  const ctx = Xrm.Utility.getGlobalContext();
  const baseUrl = ctx.getClientUrl().replace(/\\/$/, "");
  const webApi = `${{baseUrl}}/api/data/v9.2`;
  const headers = {{
    Accept: "application/json",
    "Content-Type": "application/json; charset=utf-8",
    "OData-MaxVersion": "4.0",
    "OData-Version": "4.0",
    Prefer: `odata.include-annotations="OData.Community.Display.V1.FormattedValue",odata.maxpagesize=${{PAGE_SIZE}}`
  }};

  const vehicleByPlate = new Map();
  const vehicleById = new Map();
  const driverNameById = new Map();
  const actions = [];
  const simpleList = [];
  const issues = [];
  const summary = {{
    environment: baseUrl,
    dryRun: DRY_RUN,
    servicesInScope: SERVICES.length,
    scannedServices: 0,
    agendaDriverMismatch: 0,
    resolvedVehicles: 0,
    alreadyCorrect: 0,
    dryRunWouldFillMissingVehicle: 0,
    missingVehiclesFilled: 0,
    divergenceAlerts: 0,
    skippedServiceNumberMismatch: 0,
    skippedNoLiveDriver: 0,
    skippedLiveDriverNotInTracker: 0,
    skippedOutsideTrackerCoverage: 0,
    skippedNoTrackerWindow: 0,
    skippedAmbiguousTrackerWindow: 0,
    skippedVehicleNotFound: 0,
    skippedVehicleDuplicate: 0,
    errors: 0
  }};

  function cleanGuid(value) {{
    return String(value || "").replace(/[{{}}]/g, "").trim().toLowerCase();
  }}

  function normalizeText(value) {{
    return String(value || "")
      .normalize("NFD")
      .replace(/[\\u0300-\\u036f]/g, "")
      .toLowerCase()
      .replace(/\\s+/g, " ")
      .trim();
  }}

  function escapeODataText(value) {{
    return String(value || "").replace(/'/g, "''");
  }}

  function parseLocalDate(value) {{
    if (!value) return null;
    if (value instanceof Date) return Number.isNaN(value.getTime()) ? null : value;
    const text = String(value).trim();
    const match = /^(\\d{{4}})-(\\d{{2}})-(\\d{{2}})[ T](\\d{{2}}):(\\d{{2}}):(\\d{{2}})$/.exec(text);
    if (match) {{
      return new Date(
        Number(match[1]),
        Number(match[2]) - 1,
        Number(match[3]),
        Number(match[4]),
        Number(match[5]),
        Number(match[6])
      );
    }}
    const parsed = new Date(text);
    return Number.isNaN(parsed.getTime()) ? null : parsed;
  }}

  function formatLocalDate(date) {{
    if (!(date instanceof Date) || Number.isNaN(date.getTime())) return "";
    const pad = (value) => String(value).padStart(2, "0");
    return `${{date.getFullYear()}}-${{pad(date.getMonth() + 1)}}-${{pad(date.getDate())}} ${{pad(date.getHours())}}:${{pad(date.getMinutes())}}:${{pad(date.getSeconds())}}`;
  }}

  function addAction(action) {{
    if (actions.length < ACTION_LOG_LIMIT) actions.push(action);
  }}

  function addIssue(kind, entry, extra = {{}}) {{
    issues.push({{ kind, serviceNumber: entry.serviceNumber, serviceId: entry.serviceId, ...extra }});
  }}

  function addSimpleRow(serviceNumber, serviceDate, liveDriverName, oldPlate, newPlate) {{
    simpleList.push({{
      ID: String(serviceNumber || "").trim(),
      Data: String(serviceDate || "").trim(),
      Motorista: String(liveDriverName || "").trim(),
      "Placa antiga": String(oldPlate || "").trim().toUpperCase(),
      "Placa Nova": String(newPlate || "").trim().toUpperCase()
    }});
  }}

  async function request(method, pathOrUrl, body) {{
    const url = /^https?:\\/\\//i.test(pathOrUrl) ? pathOrUrl : `${{webApi}}${{pathOrUrl}}`;
    const response = await fetch(url, {{
      method,
      headers,
      credentials: "same-origin",
      body: body ? JSON.stringify(body) : undefined
    }});
    const text = await response.text();
    if (!response.ok) throw new Error(`${{method}} ${{url}}\\n${{response.status}} ${{response.statusText}}\\n${{text}}`);
    return text ? JSON.parse(text) : null;
  }}

  function prepareTrackerWindows() {{
    for (const windows of Object.values(TRACKER_WINDOWS_BY_DRIVER)) {{
      for (const window of windows) {{
        window.startMs = parseLocalDate(window.start)?.getTime() ?? NaN;
        window.endMs = parseLocalDate(window.end)?.getTime() ?? NaN;
      }}
      windows.sort((a, b) => a.startMs - b.startMs);
    }}
  }}

  function findTrackerWindow(driverKey, serviceDate) {{
    const windows = TRACKER_WINDOWS_BY_DRIVER[driverKey] || [];
    if (!windows.length) return {{ status: "driver_not_in_tracker" }};

    const serviceMs = serviceDate.getTime();
    const hits = windows.filter((window) => window.startMs <= serviceMs && serviceMs < window.endMs);
    if (!hits.length) return {{ status: "no_window" }};

    const plates = [...new Set(hits.map((window) => window.plate))].sort();
    if (plates.length !== 1) return {{ status: "ambiguous", hits, plates }};

    return {{ status: "ok", window: hits[0], plate: plates[0] }};
  }}

  async function resolveVehicleByPlate(plate) {{
    const key = String(plate || "").trim().toUpperCase();
    if (vehicleByPlate.has(key)) return vehicleByPlate.get(key);
    const result = await request(
      "GET",
      `/${{VEHICLE_ENTITY_SET}}?$select=${{VEHICLE_ID_FIELD}},${{VEHICLE_PLATE_FIELD}}&$filter=${{VEHICLE_PLATE_FIELD}} eq '${{escapeODataText(key)}}'&$top=2`
    );
    const rows = result?.value || [];
    let resolved;
    if (rows.length === 1) {{
      resolved = {{ ok: true, id: cleanGuid(rows[0][VEHICLE_ID_FIELD]), plate: rows[0][VEHICLE_PLATE_FIELD] }};
      summary.resolvedVehicles += 1;
    }} else if (rows.length === 0) {{
      resolved = {{ ok: false, reason: "not_found" }};
    }} else {{
      resolved = {{ ok: false, reason: "duplicate", count: rows.length }};
    }}
    vehicleByPlate.set(key, resolved);
    return resolved;
  }}

  async function resolveVehicleById(vehicleId) {{
    const key = cleanGuid(vehicleId);
    if (!key) return {{ id: "", plate: "" }};
    if (vehicleById.has(key)) return vehicleById.get(key);
    const record = await request(
      "GET",
      `/${{VEHICLE_ENTITY_SET}}(${{key}})?$select=${{VEHICLE_ID_FIELD}},${{VEHICLE_PLATE_FIELD}}`
    );
    const resolved = {{
      id: cleanGuid(record?.[VEHICLE_ID_FIELD]),
      plate: String(record?.[VEHICLE_PLATE_FIELD] || "").trim().toUpperCase()
    }};
    vehicleById.set(key, resolved);
    return resolved;
  }}

  async function readService(entry) {{
    const id = cleanGuid(entry.serviceId);
    return request(
      "GET",
      `/${{SERVICE_ENTITY_SET}}(${{id}})?$select=${{SERVICE_ID_FIELD}},${{SERVICE_NUMBER_FIELD}},${{SERVICE_DATE_FIELD}},${{SERVICE_VEHICLE_LOOKUP}},${{SERVICE_DRIVER_LOOKUP}}`
    );
  }}

  async function resolveDriverName(service) {{
    const formatted = String(service?.[DRIVER_FORMATTED] || "").trim();
    if (formatted) return formatted;

    const driverId = cleanGuid(service?.[SERVICE_DRIVER_LOOKUP]);
    if (!driverId) return "";
    if (driverNameById.has(driverId)) return driverNameById.get(driverId);

    const employee = await request(
      "GET",
      `/${{EMPLOYEE_ENTITY_SET}}(${{driverId}})?$select=${{EMPLOYEE_ID_FIELD}},${{EMPLOYEE_NAME_FIELD}}`
    );
    const name = String(employee?.[EMPLOYEE_NAME_FIELD] || "").trim();
    driverNameById.set(driverId, name);
    return name;
  }}

  function parseServiceDate(service, entry) {{
    const rawDate = service?.[SERVICE_DATE_FIELD];
    const parsed = rawDate ? new Date(rawDate) : null;
    if (parsed instanceof Date && !Number.isNaN(parsed.getTime())) return parsed;
    return parseLocalDate(entry.startLocalFromAgenda);
  }}

  prepareTrackerWindows();
  const coverageStart = parseLocalDate(TRACKER_COVERAGE.start);
  const coverageEnd = parseLocalDate(TRACKER_COVERAGE.end);
  let processedServices = 0;

  for (const entry of SERVICES) {{
    try {{
      processedServices += 1;
      if (processedServices === 1 || processedServices % PROGRESS_EVERY === 0 || processedServices === SERVICES.length) {{
        console.log(`[service-vehicle-correction] progresso ${{processedServices}}/${{SERVICES.length}}`);
      }}
      summary.scannedServices += 1;
      const service = await readService(entry);
      const currentNumber = String(service?.[SERVICE_NUMBER_FIELD] || "").trim();
      if (STRICT_SERVICE_NUMBER && currentNumber !== entry.serviceNumber) {{
        summary.skippedServiceNumberMismatch += 1;
        addIssue("service_number_mismatch", entry, {{ currentNumber }});
        continue;
      }}

      const serviceDate = parseServiceDate(service, entry);
      if (!serviceDate) {{
        addIssue("service_date_missing", entry);
        continue;
      }}
      if (coverageStart && coverageEnd && (serviceDate < coverageStart || serviceDate > coverageEnd)) {{
        summary.skippedOutsideTrackerCoverage += 1;
        continue;
      }}

      const liveDriverName = await resolveDriverName(service);
      const liveDriverKey = normalizeText(liveDriverName);
      if (!liveDriverKey) {{
        summary.skippedNoLiveDriver += 1;
        addIssue("live_driver_missing", entry);
        continue;
      }}

      if (normalizeText(entry.driverNameFromAgenda) && normalizeText(entry.driverNameFromAgenda) !== liveDriverKey) {{
        summary.agendaDriverMismatch += 1;
      }}

      const match = findTrackerWindow(liveDriverKey, serviceDate);
      if (match.status === "driver_not_in_tracker") {{
        summary.skippedLiveDriverNotInTracker += 1;
        continue;
      }}
      if (match.status === "no_window") {{
        summary.skippedNoTrackerWindow += 1;
        continue;
      }}
      if (match.status === "ambiguous") {{
        summary.skippedAmbiguousTrackerWindow += 1;
        addIssue("ambiguous_tracker_window", entry, {{
          liveDriverName,
          serviceDate: formatLocalDate(serviceDate),
          plates: match.plates
        }});
        continue;
      }}

      const vehicle = await resolveVehicleByPlate(match.plate);
      if (!vehicle.ok) {{
        if (vehicle.reason === "duplicate") summary.skippedVehicleDuplicate += 1;
        else summary.skippedVehicleNotFound += 1;
        addIssue(`vehicle_${{vehicle.reason}}`, entry, {{ targetPlate: match.plate, count: vehicle.count || 0 }});
        continue;
      }}

      const currentVehicleId = cleanGuid(service?.[SERVICE_VEHICLE_LOOKUP]);
      const currentVehicle = await resolveVehicleById(currentVehicleId);
      if (currentVehicleId === vehicle.id) {{
        summary.alreadyCorrect += 1;
        addAction({{
          action: "already_correct",
          serviceNumber: entry.serviceNumber,
          liveDriverName,
          targetPlate: match.plate,
          serviceDate: formatLocalDate(serviceDate)
        }});
        continue;
      }}

      if (currentVehicleId) {{
        summary.divergenceAlerts += 1;
        addSimpleRow(
          entry.serviceNumber,
          formatLocalDate(serviceDate),
          liveDriverName,
          currentVehicle.plate || service?.[VEHICLE_FORMATTED] || "",
          match.plate
        );
        addAction({{
          action: "divergence_only",
          serviceNumber: entry.serviceNumber,
          serviceId: entry.serviceId,
          liveDriverName,
          currentVehicle: currentVehicle.plate || service?.[VEHICLE_FORMATTED] || "",
          trackerVehicle: match.plate,
          serviceDate: formatLocalDate(serviceDate),
          trackerWindowStart: match.window.start,
          trackerWindowEnd: match.window.end
        }});
        continue;
      }}

      const action = {{
        action: DRY_RUN ? "would_fill_missing_vehicle" : "filled_missing_vehicle",
        serviceNumber: entry.serviceNumber,
        serviceId: entry.serviceId,
        liveDriverName,
        agendaDriverName: entry.driverNameFromAgenda,
        currentVehicle: service?.[VEHICLE_FORMATTED] || "",
        agendaVehiclePlate: entry.vehiclePlateFromAgenda || "",
        targetPlate: match.plate,
        serviceDate: formatLocalDate(serviceDate),
        serviceDateFormatted: service?.[DATE_FORMATTED] || "",
        trackerWindowStart: match.window.start,
        trackerWindowEnd: match.window.end
      }};

      if (DRY_RUN) {{
        summary.dryRunWouldFillMissingVehicle += 1;
        addSimpleRow(
          entry.serviceNumber,
          formatLocalDate(serviceDate),
          liveDriverName,
          service?.[VEHICLE_FORMATTED] || "",
          match.plate
        );
        addAction(action);
        continue;
      }}

      await request("PATCH", `/${{SERVICE_ENTITY_SET}}(${{cleanGuid(entry.serviceId)}})`, {{
        [`${{SERVICE_VEHICLE_NAV}}@odata.bind`]: `/${{VEHICLE_ENTITY_SET}}(${{vehicle.id}})`
      }});
      summary.missingVehiclesFilled += 1;
      addSimpleRow(
        entry.serviceNumber,
        formatLocalDate(serviceDate),
        liveDriverName,
        service?.[VEHICLE_FORMATTED] || "",
        match.plate
      );
      addAction(action);
    }} catch (error) {{
      summary.errors += 1;
      addIssue("error", entry, {{ message: error?.message || String(error) }});
      console.error("[service-vehicle-correction] erro", entry.serviceNumber, error);
    }}
  }}

  simpleList.sort((a, b) => a.Data.localeCompare(b.Data) || a.ID.localeCompare(b.ID));
  const result = {{ summary, lista: simpleList, actions, issues }};
  window.ServiceVehicleCorrectionLastResult = result;
  console.log("[service-vehicle-correction] resumo", summary);
  console.table(simpleList);
  if (issues.length) console.table(issues.slice(0, 80));
  return result;
}})();
'''
    OUTPUT_SCRIPT.write_text(script, encoding="utf-8")


def main():
    tracker_windows_by_driver, tracker_summary = load_tracker_windows()
    services, service_summary = load_service_scope()
    summary = {
        "generatedAtLocal": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "agendaFile": str(AGENDA_FILE),
        "trackerFiles": [str(path) for path in TRACKER_FILES],
        "matchRule": (
            "Console script reads live Dataverse driver/date per service, then matches "
            "the live driver against tracker vehicle windows. Agenda driver is audit only."
        ),
        **tracker_summary,
        **service_summary,
    }
    write_json_and_csv(summary, services, tracker_windows_by_driver)
    write_console_script(services, tracker_windows_by_driver, tracker_summary)
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
