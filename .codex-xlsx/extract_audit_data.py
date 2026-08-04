import json
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook

agenda_path, pagantes_path, output_path = sys.argv[1:4]

def clean(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat(sep=" ")
    return value

def money(value):
    if value in (None, ""):
        return 0.0
    try:
        return float(Decimal(str(value)))
    except InvalidOperation:
        raise ValueError(f"Valor monetário inválido: {value!r}")

def op_key(value):
    return " ".join(str(value or "").strip().upper().split())

def records(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = next(ws for ws in wb.worksheets if ws.max_row and ws.max_column)
    rows = ws.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    for excel_row, values in enumerate(rows, start=2):
        yield excel_row, dict(zip(headers, values))

services = []
for excel_row, row in records(agenda_path):
    op = op_key(row.get("OP"))
    if not op:
        continue
    services.append({
        "source_row": excel_row,
        "op": op,
        "id": clean(row.get("ID")),
        "date": clean(row.get("Data e horário de saída")),
        "client": clean(row.get("Cliente")),
        "passenger": clean(row.get("Pax - VIEW")),
        "operation_status": clean(row.get("Status de Operação")),
        "billing_status": clean(row.get("Status de Faturamento")),
        "value": money(row.get("Valor Total (Composição de Preço) (Composição de Preços)")),
        "base_value": money(row.get("Valor Total (Base) (Composição de Preço) (Composição de Preços)")),
    })

payers = []
for excel_row, row in records(pagantes_path):
    op = op_key(row.get("Ordem de Pagamento"))
    if not op:
        continue
    payers.append({
        "source_row": excel_row,
        "op": op,
        "id": clean(row.get("ID")),
        "payer": clean(row.get("Pagante")),
        "payment_status": clean(row.get("Status do Pagamento")),
        "payment_method": clean(row.get("Forma de Pagamento")),
        "value": money(row.get("Valor")),
        "created": clean(row.get("Data de Criação")),
    })

payload = {
    "sources": {"agenda": agenda_path, "pagantes": pagantes_path},
    "services": services,
    "payers": payers,
}
with open(output_path, "w", encoding="utf-8") as handle:
    json.dump(payload, handle, ensure_ascii=False, indent=2)

service_totals = {}
for row in services:
    item = service_totals.setdefault(row["op"], {"count": 0, "value": 0.0})
    item["count"] += 1
    item["value"] += row["value"]

payer_totals = {}
for row in payers:
    item = payer_totals.setdefault(row["op"], {"count": 0, "all": 0.0, "active": 0.0, "cancelled": 0.0})
    item["count"] += 1
    item["all"] += row["value"]
    if str(row["payment_status"] or "").strip().casefold() == "cancelado":
        item["cancelled"] += row["value"]
    else:
        item["active"] += row["value"]

ops = sorted(set(service_totals) | set(payer_totals))
summary = []
for op in ops:
    s = service_totals.get(op, {"count": 0, "value": 0.0})
    p = payer_totals.get(op, {"count": 0, "all": 0.0, "active": 0.0, "cancelled": 0.0})
    diff = round(p["active"] - s["value"], 2)
    if not s["count"]:
        status = "SEM_SERVIÇO_NA_GERAL"
    elif not p["count"]:
        status = "SEM_PAGANTE"
    elif diff == 0:
        status = "BATE"
    elif diff < 0:
        status = "PAGANTES_A_MENOS"
    else:
        status = "PAGANTES_A_MAIS"
    summary.append({"op": op, "services": s["count"], "service_total": round(s["value"], 2), "payers": p["count"], "payer_total_all": round(p["all"], 2), "payer_total_active": round(p["active"], 2), "cancelled": round(p["cancelled"], 2), "difference": diff, "status": status})

print(json.dumps({
    "service_rows": len(services),
    "payer_rows": len(payers),
    "ops": len(ops),
    "status_counts": {status: sum(1 for row in summary if row["status"] == status) for status in sorted({row["status"] for row in summary})},
    "differences": [row for row in summary if row["status"] != "BATE"],
}, ensure_ascii=False))
