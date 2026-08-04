import json
import sys
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

for path in sys.argv[1:]:
    wb = load_workbook(path, read_only=True, data_only=False)
    print(json.dumps({"path": path, "sheets": wb.sheetnames}, ensure_ascii=False))
    for ws in wb.worksheets:
        if not ws.max_row or not ws.max_column:
            print(json.dumps({"sheet": ws.title, "max_row": ws.max_row, "max_column": ws.max_column, "rows": []}, ensure_ascii=False))
            continue
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True):
            rows.append(list(row[:52]))
        print(json.dumps({"sheet": ws.title, "max_row": ws.max_row, "max_column": ws.max_column, "rows": rows}, ensure_ascii=False, default=str))
